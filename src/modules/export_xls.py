from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

def export_xls(data, file_xls):
    # Исходные данные

    # Создаем Excel-файл
    wb = Workbook()
    ws = wb.active

    # Устанавливаем заголовок листа
    ws.title = "Специалисты"

    # Заголовки первой строки
    header_row1 = [
        "№ п/п", "Фамилия И.О.", "Подразделение", "Должность", "Звание", "Образование",
        "Приказ о допуске", "Результаты", "", "", "", "Допуск до"
    ]

    # Заголовки второй строки
    header_row2 = [
        "", "", "", "", "", "", "", "Теор", "Физо", "Тех", "Соб", ""
    ]

    # Заполняем первую и вторую строки
    ws.append(header_row1)
    ws.append(header_row2)

    # Объединяем ячейки для колонок с 8 по 11 в первой строке
    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=11)

    # Объединяем ячейки в колонках 1, 2, 3, 4, 5, 6, 7, 12 для строк 1 и 2
    columns_to_merge = [1, 2, 3, 4, 5, 6, 7, 12]
    for col in columns_to_merge:
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    # Выравнивание текста по центру и перенос текста для всех ячеек заголовков
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=12):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Создаем стиль обводки
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Применяем обводку ко всем ячейкам заголовков
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=12):
        for cell in row:
            cell.border = thin_border

    # Устанавливаем высоту первой строки
    #ws.row_dimensions[1].height = 65  # Устанавливаем высоту в 65 пунктов        

    # Добавляем данные и обводим ячейки
    for row_index, row_data in enumerate(data, start=3):  # Начинаем с 2, т.к. 1 - заголовок
        for col_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.value = value
            cell.border = thin_border
            
            # Выравнивание данных по центру для колонок F и H
            if col_index in [1, 6, 7, 8, 9, 10, 11, 12 ]:  # F = 6
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Настройка ширины колонок F и G
    ws.column_dimensions["A"].width = 5  
    ws.column_dimensions["B"].width = 35  
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 30  
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 5          
    ws.column_dimensions["I"].width = 5  
    ws.column_dimensions["J"].width = 5  
    ws.column_dimensions["K"].width = 5  
    ws.column_dimensions["L"].width = 10  

    # Сохраняем файл

    wb.save(file_xls)


